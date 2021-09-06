import excel2img
from openpyxl import load_workbook

#used for converting xlsx files to png
def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames



for i,val in enumerate(get_sheetnames_xlsx("РКМ_2021г.xlsx")):
    print(val)
    excel2img.export_img("РКМ_2021г.xlsx", f'train2\\000{i}.png', val, None)
